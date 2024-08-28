import openpyxl
import requests
from io import BytesIO
import streamlit as st

# Set page configuration
st.set_page_config(page_title="Network Level Impact Calculator", layout="wide")

# Add custom CSS to make it look darker
st.markdown(
    """
    <style>
    .reportview-container {
        background: #2e2e2e;
        color: #ffffff;
    }
    .sidebar .sidebar-content {
        background: #1e1e1e;
        color: #ffffff;
    }
    .widget-label {
        color: #ffffff;
    }
    .stButton > button {
        background-color: #555555;
        color: #ffffff;
        border: none;
        border-radius: 5px;
        padding: 10px 20px;
    }
    .stButton > button:hover {
        background-color: #777777;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# Raw URL of the file on GitHub
file_url = "https://github.com/ShadSpidy/Network-Level-Impact-2.0/raw/main/drc_npi.xlsx"

def download_file(url):
    response = requests.get(url)
    response.raise_for_status()
    return BytesIO(response.content)

try:
    # Load the Excel workbook from GitHub
    response = download_file(file_url)
    npi_file = openpyxl.load_workbook(response)
    npi_report = npi_file["Sheet"]

    # Extract node-wise traffic data
    node_wise_traffic = {}
    for row in range(118, 134):
        node_name = npi_report.cell(row, 3).value.upper().replace(' ', '')
        traffic = npi_report.cell(row, 7).value
        node_wise_traffic[node_name] = traffic

    total_traffic = sum(node_wise_traffic.values())

    # Streamlit app
    st.title("Network Level Impact Calculator")

    # User input for impacted nodes
    user_input = st.text_input("Please enter the impacted node(s) (separate multiple nodes with commas):").upper()

    if user_input:  # Check if user_input is not empty
        impacted_nodes = [node.strip() for node in user_input.split(',')]
        impacted_node_traffic = sum(node_wise_traffic.get(node, 0) for node in impacted_nodes)

        if impacted_node_traffic == 0:
            st.error("Invalid input. Please enter valid node names.")
        else:
            # User input for current percentage impact
            percentage_impact = st.number_input("Enter the current percentage impact (%):", min_value=0.0, max_value=100.0, step=0.1)

            # Calculate network level impact
            NW_level_impact = ((impacted_node_traffic / total_traffic) * 100) * (percentage_impact / 100)
            st.markdown(f"<h2 style='font-size:24px;'>Current Network Level Impact is: {round(NW_level_impact, 2)}%</h3>", unsafe_allow_html=True)

    else:
        st.info("Please enter the impacted node(s) to get results.")

except Exception as e:
    st.error(f"An error occurred: {e}")

# credit
st.markdown("**Created by: Shadman**", unsafe_allow_html=True)
